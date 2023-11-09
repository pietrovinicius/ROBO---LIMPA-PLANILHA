#21/10/2023
#@PLima
#ROBO - LIMPA PLANILHA
import openpyxl
import string
import unicodedata


def remover_acentos_e_caracteres_especiais(texto):
    caracteres_especiais = [
        "!", "\"", "#", "$", "%", "&", "'", "*", "+", ",", "-",
        ".", "/", ":", ";", "<", "=", ">", "?", "@", "[", "]", "^", "_",
        "`", "{", "|", "}", "~"
    ]   
    for caractere in caracteres_especiais:
        texto = texto.replace(caractere, "")     
        texto = unicodedata.normalize("NFKD", texto)
        texto = texto.encode("ascii", "ignore").decode("utf-8")
        texto = texto.upper()
        #print(f"remover_acentos_e_caracteres_especiais: {texto}")
    return texto        

try:
    print("================================= INICIALIZADO ======================")
    
    # Abre o arquivo Excel
    print("Abre o arquivo Excel;")
    wb = openpyxl.load_workbook("planilha.xlsx")

    # Seleciona a aba da planilha
    print("Seleciona a aba da planilha")
    sheet = wb["aba"]
    
    # Cria uma lista com os nomes da coluna A
    print("Cria uma lista com os nomes das 4 colunas;")
    coluna_1 = []
    coluna_2 = []
    coluna_3 = []
    coluna_4 = []
    for i in range(1, sheet.max_row + 1):
        #print(sheet["A" + str(i)].value)
        coluna_1.append(remover_acentos_e_caracteres_especiais(sheet["A" + str(i)].value))  
        coluna_2.append(remover_acentos_e_caracteres_especiais(sheet["B" + str(i)].value))  
        coluna_3.append(remover_acentos_e_caracteres_especiais(sheet["C" + str(i)].value))  
        coluna_4.append(remover_acentos_e_caracteres_especiais(sheet["D" + str(i)].value))          

    # Salva os nomes sem caracteres especiais na planilha
    print("Salva os nomes sem caracteres especiais na planilha;\n")
    for i in range(1, sheet.max_row + 1):
        #print( nomes_limpos[i - 1])
        sheet["A" + str(i)].value = coluna_1[i - 1]
        sheet["B" + str(i)].value = coluna_2[i - 1]
        sheet["C" + str(i)].value = coluna_3[i - 1]
        sheet["D" + str(i)].value = coluna_4[i - 1]
        #print(coluna_1[i - 1] + " - " + coluna_2[i - 1])
    
    # Salva o arquivo Excel
    wb.save("planilha_limpa_.xlsx")
    print("wb.save(planilha_limpa_.xlsx);")
    print(f"Planilha: planilha.xlsx\nAba: {wb.sheetnames}")
    print("================================= FINALIZADO ======================")
except Exception as erro:
    print(f"Error: {erro}")
